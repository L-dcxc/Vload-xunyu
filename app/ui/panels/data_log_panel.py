from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Any

from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtWidgets import (
    QApplication,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QProgressDialog,
    QPushButton,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)


class DataLogPanel(QFrame):
    data_imported = pyqtSignal(list)  # 导入数据信号，参数为数据列表 [(t, v, i, p, r), ...]
    export_requested = pyqtSignal(str, str)  # 请求导出信号 (file_path, selected_filter)
    clear_requested = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setProperty("card", "true")

        root = QVBoxLayout(self)
        root.setContentsMargins(14, 12, 14, 12)
        root.setSpacing(10)

        title = QLabel("数据与日志")
        title.setStyleSheet("font-size: 14px; font-weight: 800;")

        self.tabs = QTabWidget()

        data_tab = QWidget()
        data_layout = QVBoxLayout(data_tab)
        data_layout.setContentsMargins(0, 0, 0, 0)
        data_layout.setSpacing(10)

        export_row = QHBoxLayout()
        self.btn_import = QPushButton("导入波形")
        self.btn_import.setProperty("variant", "secondary")
        self.btn_export = QPushButton("导出数据")
        self.btn_export.setProperty("variant", "secondary")
        self.btn_clear_data = QPushButton("清空数据")
        self.btn_clear_data.setProperty("variant", "secondary")
        export_row.addWidget(self.btn_import)
        export_row.addWidget(self.btn_export)
        export_row.addWidget(self.btn_clear_data)
        export_row.addStretch(1)

        self.data_table = QTableWidget(0, 5)
        self.data_table.setHorizontalHeaderLabels(["时间", "电压(V)", "电流(A)", "功率(W)", "内阻(Ω)"])
        self.data_table.verticalHeader().setVisible(False)
        self.data_table.setAlternatingRowColors(True)
        self.data_table.setSelectionBehavior(self.data_table.SelectionBehavior.SelectRows)
        self.data_table.setEditTriggers(self.data_table.EditTrigger.NoEditTriggers)
        self.data_table.horizontalHeader().setStretchLastSection(True)
        self.data_table.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignLeft)

        data_layout.addLayout(export_row)
        data_layout.addWidget(self.data_table, 1)

        run_log_tab = QWidget()
        run_log_layout = QVBoxLayout(run_log_tab)
        run_log_layout.setContentsMargins(0, 0, 0, 0)
        run_log_layout.setSpacing(10)

        self.run_log = QTextEdit()
        self.run_log.setReadOnly(True)
        self.run_log.setPlaceholderText("运行日志将显示在这里…")

        run_log_layout.addWidget(self.run_log, 1)

        comm_log_tab = QWidget()
        comm_log_layout = QVBoxLayout(comm_log_tab)
        comm_log_layout.setContentsMargins(0, 0, 0, 0)
        comm_log_layout.setSpacing(10)

        comm_row = QHBoxLayout()
        self.btn_clear_comm = QPushButton("清空通信日志")
        self.btn_clear_comm.setProperty("variant", "secondary")
        comm_row.addWidget(self.btn_clear_comm)
        comm_row.addStretch(1)

        self.comm_log = QTextEdit()
        self.comm_log.setReadOnly(True)
        self.comm_log.setPlaceholderText("通信日志将显示在这里…")

        comm_log_layout.addLayout(comm_row)
        comm_log_layout.addWidget(self.comm_log, 1)

        term_tab = QWidget()
        term_layout = QVBoxLayout(term_tab)
        term_layout.setContentsMargins(0, 0, 0, 0)
        term_layout.setSpacing(10)

        self.term_output = QTextEdit()
        self.term_output.setReadOnly(True)
        self.term_output.setPlaceholderText("SCPI 终端输出…")

        input_row = QHBoxLayout()
        self.term_input = QLineEdit()
        self.term_input.setPlaceholderText("输入 SCPI 命令，例如 *IDN?")
        self.btn_term_send = QPushButton("发送")

        input_row.addWidget(self.term_input, 1)
        input_row.addWidget(self.btn_term_send)

        term_layout.addWidget(self.term_output, 1)
        term_layout.addLayout(input_row)

        self.tabs.addTab(data_tab, "数据")
        self.tabs.addTab(run_log_tab, "运行日志")
        self.tabs.addTab(comm_log_tab, "通信日志")
        self.tabs.addTab(term_tab, "SCPI 终端")

        root.addWidget(title)
        root.addWidget(self.tabs, 1)

        self.btn_clear_comm.clicked.connect(self.comm_log.clear)
        self.btn_import.clicked.connect(self._on_import_clicked)
        self.btn_export.clicked.connect(self._on_export_clicked)
        self.btn_clear_data.clicked.connect(self.clear_requested)

    def clear_data(self) -> None:
        self.data_table.setRowCount(0)

    def load_imported_data(
        self,
        rows: list[tuple[float, float, float, float, float]],
        progress: QProgressDialog | None = None,
    ) -> bool:
        self.data_table.setUpdatesEnabled(False)
        try:
            self.data_table.setRowCount(0)
            self.data_table.setRowCount(len(rows))
            for row_idx, (t_s, v, i, p, r) in enumerate(rows):
                if progress is not None and (row_idx % 200 == 0 or row_idx == len(rows) - 1):
                    progress.setValue(row_idx)
                    QApplication.processEvents()
                    if progress.wasCanceled():
                        return False
                ts = datetime.fromtimestamp(t_s).strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                for col, val in enumerate([ts, f"{v:.3f}", f"{i:.3f}", f"{p:.3f}", f"{r:.3f}"]):
                    self.data_table.setItem(row_idx, col, QTableWidgetItem(val))
        finally:
            self.data_table.setUpdatesEnabled(True)
        self.data_table.scrollToBottom()
        return True

    def data_row_count(self) -> int:
        return self.data_table.rowCount()

    def open_export_dialog(self) -> None:
        self._on_export_clicked()

    def append_run_log(self, text: str) -> None:
        ts = datetime.now().strftime("%H:%M:%S")
        self.run_log.append(f"[{ts}] {text}")

    def append_comm_log(self, text: str) -> None:
        ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        self.comm_log.append(f"[{ts}] {text}")

    def append_term(self, text: str) -> None:
        self.term_output.append(text)

    def append_data(self, v: float, i: float, p: float, r: float) -> None:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
        row = self.data_table.rowCount()
        self.data_table.insertRow(row)
        for col, val in enumerate([ts, f"{v:.3f}", f"{i:.3f}", f"{p:.3f}", f"{r:.3f}"]):
            self.data_table.setItem(row, col, QTableWidgetItem(val))

        if self.data_table.rowCount() > 50000:
            self.data_table.removeRow(0)

        self.data_table.scrollToBottom()

    def _on_import_clicked(self) -> None:
        """导入波形数据"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "导入波形数据", "", "CSV 文件 (*.csv);;Excel 文件 (*.xlsx *.xls);;所有文件 (*.*)"
        )

        if not file_path:
            return

        try:
            data_list = []
            file_ext = Path(file_path).suffix.lower()

            def _find_col(header_cells: list[str], candidates: list[str]) -> int | None:
                for c in candidates:
                    for idx, h in enumerate(header_cells):
                        if h == c:
                            return idx
                for c in candidates:
                    for idx, h in enumerate(header_cells):
                        if c in h:
                            return idx
                return None

            def _parse_ts(ts_text: str) -> datetime | None:
                s = str(ts_text).strip()
                if not s:
                    return None
                for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S"):
                    try:
                        return datetime.strptime(s, fmt)
                    except ValueError:
                        pass
                return None

            if file_ext == ".csv":
                # 读取 CSV 文件
                import csv

                with open(file_path, "r", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    header = next(reader, [])
                    header_cells = [str(x).strip() for x in header]

                    idx_ts = _find_col(header_cells, ["时间"])
                    idx_v = _find_col(header_cells, ["电压(V)", "电压"])
                    idx_i = _find_col(header_cells, ["电流(A)", "电流"])
                    idx_p = _find_col(header_cells, ["功率(W)", "功率"])
                    idx_r = _find_col(header_cells, ["内阻(Ω)", "内阻", "内阻(Ohm)"])

                    if None in (idx_v, idx_i, idx_p, idx_r):
                        QMessageBox.warning(self, "格式错误", "文件缺少必要列（电压/电流/功率/内阻），无法导入")
                        return

                    base_dt: datetime | None = None

                    for row in reader:
                        if not row:
                            continue
                        try:
                            v = float(row[idx_v])
                            i = float(row[idx_i])
                            p = float(row[idx_p])
                            r = float(row[idx_r])

                            if idx_ts is not None and idx_ts < len(row):
                                dt = _parse_ts(row[idx_ts])
                            else:
                                dt = None

                            if dt is not None:
                                if base_dt is None:
                                    base_dt = dt
                                t = (dt - base_dt).total_seconds()
                            else:
                                t = len(data_list) * 0.2

                            data_list.append((t, v, i, p, r))
                        except (ValueError, IndexError):
                            continue

            elif file_ext in [".xlsx", ".xls"]:
                # 读取 Excel 文件
                try:
                    import openpyxl

                    wb = openpyxl.load_workbook(file_path)
                    ws = wb.active

                    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                    if not header_row:
                        QMessageBox.warning(self, "导入失败", "文件为空")
                        return

                    header_cells = [str(x).strip() for x in header_row]
                    idx_ts = _find_col(header_cells, ["时间"])
                    idx_v = _find_col(header_cells, ["电压(V)", "电压"])
                    idx_i = _find_col(header_cells, ["电流(A)", "电流"])
                    idx_p = _find_col(header_cells, ["功率(W)", "功率"])
                    idx_r = _find_col(header_cells, ["内阻(Ω)", "内阻", "内阻(Ohm)"])

                    if None in (idx_v, idx_i, idx_p, idx_r):
                        QMessageBox.warning(self, "格式错误", "文件缺少必要列（电压/电流/功率/内阻），无法导入")
                        return

                    base_dt: datetime | None = None

                    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                        if not row:
                            continue
                        try:
                            v = float(row[idx_v])
                            i = float(row[idx_i])
                            p = float(row[idx_p])
                            r = float(row[idx_r])

                            if idx_ts is not None and idx_ts < len(row):
                                dt = _parse_ts(row[idx_ts])
                            else:
                                dt = None

                            if dt is not None:
                                if base_dt is None:
                                    base_dt = dt
                                t = (dt - base_dt).total_seconds()
                            else:
                                t = idx * 0.2

                            data_list.append((t, v, i, p, r))
                        except (ValueError, TypeError, IndexError):
                            continue
                except ImportError:
                    QMessageBox.warning(
                        self, "缺少依赖", "读取 Excel 文件需要安装 openpyxl 库\n请运行: pip install openpyxl"
                    )
                    return
            else:
                QMessageBox.warning(self, "不支持的格式", "仅支持 CSV 和 Excel 文件")
                return

            if not data_list:
                QMessageBox.warning(self, "导入失败", "文件中没有有效数据")
                return

            # 发射信号，通知主窗口更新波形
            self.data_imported.emit(data_list)
            QMessageBox.information(self, "导入成功", f"成功导入 {len(data_list)} 条数据")

        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"导入数据时发生错误：{str(e)}")

    def _on_export_clicked(self) -> None:
        """导出数据"""
        if self.data_table.rowCount() == 0:
            QMessageBox.warning(self, "无数据", "没有可导出的数据")
            return

        # 生成默认文件名（带日期时间）
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"VLoad_数据_{timestamp}.csv"

        file_path, selected_filter = QFileDialog.getSaveFileName(
            self, "导出数据", default_name, "CSV 文件 (*.csv);;Excel 文件 (*.xlsx);;所有文件 (*.*)"
        )

        if not file_path:
            return

        self.export_requested.emit(file_path, selected_filter)

    def export_with_metadata(self, file_path: str, selected_filter: str, meta: dict) -> None:
        """使用设备查询到的元数据导出"""
        mode = str(meta.get("mode", ""))
        max_i = meta.get("max_current", "")
        max_p = meta.get("max_power", "")

        try:
            file_ext = Path(file_path).suffix.lower()
            headers = [
                "时间",
                "工作模式",
                "最大电流(A)",
                "最大功率(W)",
                "电压(V)",
                "电流(A)",
                "功率(W)",
                "内阻(Ω)",
            ]

            def iter_rows():
                for row in range(self.data_table.rowCount()):
                    ts = self.data_table.item(row, 0).text() if self.data_table.item(row, 0) else ""
                    v = self.data_table.item(row, 1).text() if self.data_table.item(row, 1) else ""
                    i = self.data_table.item(row, 2).text() if self.data_table.item(row, 2) else ""
                    p = self.data_table.item(row, 3).text() if self.data_table.item(row, 3) else ""
                    r = self.data_table.item(row, 4).text() if self.data_table.item(row, 4) else ""
                    yield [ts, mode, max_i, max_p, v, i, p, r]

            if file_ext == ".csv" or "CSV" in selected_filter:
                if not file_path.endswith(".csv"):
                    file_path += ".csv"

                import csv

                with open(file_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    for row_data in iter_rows():
                        writer.writerow(row_data)

            elif file_ext == ".xlsx" or "Excel" in selected_filter:
                if not file_path.endswith(".xlsx"):
                    file_path += ".xlsx"

                try:
                    import openpyxl
                except ImportError:
                    QMessageBox.warning(
                        self, "缺少依赖", "导出 Excel 文件需要安装 openpyxl 库\n请运行: pip install openpyxl"
                    )
                    return

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "测量数据"
                ws.append(headers)
                for row_data in iter_rows():
                    ws.append(row_data)
                wb.save(file_path)
            else:
                QMessageBox.warning(self, "不支持的格式", "请选择 CSV 或 Excel 格式")
                return

            QMessageBox.information(self, "导出成功", f"数据已导出到：\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出数据时发生错误：{str(e)}")

    def export_from_session(self, file_path: str, selected_filter: str, meta: dict, data_csv_path: Path) -> None:
        mode = str(meta.get("mode", ""))
        max_i = meta.get("max_current", "")
        max_p = meta.get("max_power", "")

        headers = [
            "时间",
            "工作模式",
            "最大电流(A)",
            "最大功率(W)",
            "电压(V)",
            "电流(A)",
            "功率(W)",
            "内阻(Ω)",
        ]

        try:
            file_ext = Path(file_path).suffix.lower()

            if file_ext == ".csv" or "CSV" in selected_filter:
                if not file_path.endswith(".csv"):
                    file_path += ".csv"

                with open(file_path, "w", newline="", encoding="utf-8") as out_f:
                    writer = csv.writer(out_f)
                    writer.writerow(headers)
                    with open(data_csv_path, "r", encoding="utf-8") as in_f:
                        reader = csv.reader(in_f)
                        _ = next(reader, None)
                        for row in reader:
                            if len(row) < 5:
                                continue
                            ts, v, i, p, r = row[0], row[1], row[2], row[3], row[4]
                            writer.writerow([ts, mode, max_i, max_p, v, i, p, r])

            elif file_ext == ".xlsx" or "Excel" in selected_filter:
                if not file_path.endswith(".xlsx"):
                    file_path += ".xlsx"

                try:
                    import openpyxl
                except ImportError:
                    QMessageBox.warning(
                        self, "缺少依赖", "导出 Excel 文件需要安装 openpyxl 库\n请运行: pip install openpyxl"
                    )
                    return

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "测量数据_1"
                ws.append(headers)
                row_idx = 1
                sheet_idx = 1

                def new_sheet() -> Any:
                    nonlocal ws, sheet_idx, row_idx
                    sheet_idx += 1
                    ws = wb.create_sheet(title=f"测量数据_{sheet_idx}")
                    ws.append(headers)
                    row_idx = 1
                    return ws

                with open(data_csv_path, "r", encoding="utf-8") as in_f:
                    reader = csv.reader(in_f)
                    _ = next(reader, None)
                    for row in reader:
                        if len(row) < 5:
                            continue
                        if row_idx >= 1_048_575:
                            new_sheet()
                        ts, v, i, p, r = row[0], row[1], row[2], row[3], row[4]
                        ws.append([ts, mode, max_i, max_p, v, i, p, r])
                        row_idx += 1

                wb.save(file_path)
            else:
                QMessageBox.warning(self, "不支持的格式", "请选择 CSV 或 Excel 格式")
                return

            QMessageBox.information(self, "导出成功", f"数据已导出到：\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出数据时发生错误：{str(e)}")
